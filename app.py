# app.py
# ------------------------------------------------------------
# Asistente de Peticiones de Almacenes (Streamlit)
# - 2 carritos: importado vs manual
# - Importación Excel + matching contra catálogo
# - Búsqueda por referencia / nombre / EAN
# - Grid Color x Talla con botones + / -
# - Exportación a plantilla_pedido.xlsx (Observaciones = ref_peticion)
# ------------------------------------------------------------

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date
from typing import Dict, Tuple, Optional, List

import pandas as pd
import streamlit as st

try:
    import openpyxl
except Exception:
    openpyxl = None


# ============================================================
# CONFIG
# ============================================================

st.set_page_config(
    page_title="Peticiones almacenes",
    page_icon="📦",
    layout="wide",
)

# Estilo "Joor-like" (minimal, white, clean)
st.markdown(
    """
<style>
:root { --accent: #111; }

.block-container { padding-top: 1.2rem; padding-bottom: 2.5rem; }
h1, h2, h3 { letter-spacing: -0.02em; }
div[data-testid="stMetricValue"] { color: var(--accent); }
div[data-testid="stSidebar"] { border-right: 1px solid #eee; }
hr { margin: 1.2rem 0; }

.card {
  border: 1px solid #eee;
  border-radius: 14px;
  padding: 14px 16px;
  background: #fff;
}
.badge {
  display: inline-block;
  padding: 2px 10px;
  border-radius: 999px;
  border: 1px solid #eaeaea;
  font-size: 12px;
  background: #fafafa;
}
.small { font-size: 12px; color: #666; }
.mono { font-family: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Liberation Mono", "Courier New", monospace; }

.gridwrap {
  border: 1px solid #eee;
  border-radius: 14px;
  padding: 10px 12px;
  background: #fff;
  overflow-x: auto;
}

.cellqty {
  text-align: center;
  font-weight: 600;
  padding-top: 6px;
  padding-bottom: 6px;
}

td, th { vertical-align: middle !important; }
</style>
""",
    unsafe_allow_html=True,
)

# Almacenes (lista cerrada)
ORIGIN_OPTIONS = [
    "PET Almacén Badalona",
    "ALM-CENTRAL",
]
DEST_OPTIONS = [
    "PET T002 Marbella",
    "ALM-TIENDA",
]

# Ficheros por defecto (si están en el repo)
DEFAULT_CATALOG_PATH = "catalogue.xlsx"
DEFAULT_TEMPLATE_PATH = "plantilla_pedido.xlsx"


# ============================================================
# HELPERS / DOMAIN
# ============================================================

TALLA_MAP = {
    "XXS": "XXS",
    "XS": "XS",
    "S": "S",
    "M": "M",
    "L": "L",
    "XL": "XL",
    "XXL": "XXL",
    "XXXL": "XXXL",
}

TALLA_REGEX = re.compile(r"^\s*(XXS|XS|S|M|L|XL|XXL|XXXL|[0-9]{2,3}|[0-9]{1,2}[A-Z]?)\s*$", re.I)

REF_BRACKET_REGEX = re.compile(r"\[(?P<ref>[^\]]+)\]")
ATTR_PAREN_REGEX = re.compile(r"\((?P<attrs>[^)]+)\)\s*$")


def norm_str(x: object) -> str:
    if x is None:
        return ""
    return str(x).strip()


def norm_ref(x: object) -> str:
    # Mantén tal cual (en tu catálogo suelen ser numéricas tipo "214803" pero en peticiones puede venir "T225")
    return norm_str(x)


def norm_color(x: object) -> str:
    return norm_str(x)


def norm_talla(x: object) -> str:
    s = norm_str(x)
    up = s.upper()
    return TALLA_MAP.get(up, s)


def looks_like_talla(s: str) -> bool:
    if not s:
        return False
    return bool(TALLA_REGEX.match(s.strip()))


@dataclass(frozen=True)
class VariantKey:
    ref: str
    color: str
    talla: str


def build_catalog_indexes(cat: pd.DataFrame):
    """
    Devuelve:
      - idx_exact: (ref,color,talla) -> row dict
      - idx_ref_color: (ref,color) -> list row dict
      - idx_ref_talla: (ref,talla) -> list row dict
      - idx_ref: ref -> list row dict
    """
    idx_exact: Dict[Tuple[str, str, str], dict] = {}
    idx_ref_color: Dict[Tuple[str, str], List[dict]] = {}
    idx_ref_talla: Dict[Tuple[str, str], List[dict]] = {}
    idx_ref: Dict[str, List[dict]] = {}

    for _, r in cat.iterrows():
        ref = norm_ref(r.get("Referencia", ""))
        color = norm_color(r.get("Color", ""))
        talla = norm_talla(r.get("Talla", ""))
        ean = norm_str(r.get("EAN", ""))
        nombre = norm_str(r.get("Nombre", ""))

        row = {
            "EAN": ean,
            "Referencia": ref,
            "Color": color,
            "Talla": talla,
            "Nombre": nombre,
        }

        idx_exact[(ref, color, talla)] = row
        idx_ref_color.setdefault((ref, color), []).append(row)
        idx_ref_talla.setdefault((ref, talla), []).append(row)
        idx_ref.setdefault(ref, []).append(row)

    return idx_exact, idx_ref_color, idx_ref_talla, idx_ref


def parse_petition_line(raw: str) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    """
    A partir de un texto estilo:
      "[248202] Blusa Sabhy (NEGRO MERI, S)"
      "[248202] Blusa Sabhy (S)"
      "[248202] Blusa Sabhy (NEGRO MERI)"
    Devuelve (ref, color, talla) donde color/talla pueden ser None.
    """
    if not raw:
        return None, None, None

    raw = str(raw).strip()
    mref = REF_BRACKET_REGEX.search(raw)
    if not mref:
        return None, None, None
    ref = norm_ref(mref.group("ref"))

    mattr = ATTR_PAREN_REGEX.search(raw)
    if not mattr:
        return ref, None, None

    attrs_txt = mattr.group("attrs")
    parts = [p.strip() for p in attrs_txt.split(",") if p.strip()]

    color = None
    talla = None
    if len(parts) >= 2:
        a, b = parts[0], parts[1]
        if looks_like_talla(a) and not looks_like_talla(b):
            talla = norm_talla(a)
            color = norm_color(b)
        elif looks_like_talla(b) and not looks_like_talla(a):
            color = norm_color(a)
            talla = norm_talla(b)
        else:
            color = norm_color(a)
            talla = norm_talla(b)
    elif len(parts) == 1:
        a = parts[0]
        if looks_like_talla(a):
            talla = norm_talla(a)
        else:
            color = norm_color(a)

    return ref, color, talla


def detect_qty_column(df: pd.DataFrame) -> str:
    """
    Robusto para tu Excel tipo pivot:
      - si hay >= 3 columnas: intenta col2 si suma numérica > 0
      - si no, intenta col1
    """
    cols = list(df.columns)
    if len(cols) >= 3:
        c2 = cols[2]
        s2 = pd.to_numeric(df[c2], errors="coerce").fillna(0).sum()
        if s2 > 0:
            return c2
    if len(cols) >= 2:
        return cols[1]
    return cols[0]


def read_petition_excel(file_bytes: bytes) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(file_bytes))
    cols = list(df.columns)
    if not cols:
        return pd.DataFrame(columns=["raw", "qty"])

    raw_col = cols[0]
    qty_col = detect_qty_column(df)

    out = pd.DataFrame(
        {
            "raw": df[raw_col],
            "qty": pd.to_numeric(df[qty_col], errors="coerce").fillna(0).astype(int),
        }
    )
    return out


def match_petition_to_catalog(
    petition_df: pd.DataFrame,
    idx_exact,
    idx_ref_color,
    idx_ref_talla,
    idx_ref,
):
    """
    Devuelve:
      - matched_rows: lista dict {EAN, Referencia, Nombre, Color, Talla, Cantidad, match_level}
      - pending_rows: lista dict {raw, qty, ref, color, talla, reason}
    """
    matched = []
    pending = []

    for _, r in petition_df.iterrows():
        raw = norm_str(r.get("raw", ""))
        qty = int(r.get("qty", 0))

        # Regla: obviar filas no relevantes
        if qty <= 0:
            continue
        if "[" not in raw:
            continue

        ref, color, talla = parse_petition_line(raw)
        if not ref:
            continue

        if color and talla:
            hit = idx_exact.get((ref, color, talla))
            if hit:
                matched.append({**hit, "Cantidad": qty, "match_level": "exact"})
                continue
            pending.append(
                {
                    "raw": raw,
                    "qty": qty,
                    "ref": ref,
                    "color": color,
                    "talla": talla,
                    "reason": "No existe esa variante exacta en catálogo",
                }
            )
            continue

        if color and not talla:
            hits = idx_ref_color.get((ref, color), [])
            if len(hits) == 1:
                matched.append({**hits[0], "Cantidad": qty, "match_level": "ref+color"})
                continue
            if len(hits) > 1:
                pending.append(
                    {
                        "raw": raw,
                        "qty": qty,
                        "ref": ref,
                        "color": color,
                        "talla": None,
                        "reason": "Ambiguo: múltiples tallas para ese color",
                    }
                )
                continue
            pending.append(
                {
                    "raw": raw,
                    "qty": qty,
                    "ref": ref,
                    "color": color,
                    "talla": None,
                    "reason": "No se encontró ref+color en catálogo",
                }
            )
            continue

        if talla and not color:
            hits = idx_ref_talla.get((ref, talla), [])
            if len(hits) == 1:
                matched.append({**hits[0], "Cantidad": qty, "match_level": "ref+talla"})
                continue
            if len(hits) > 1:
                pending.append(
                    {
                        "raw": raw,
                        "qty": qty,
                        "ref": ref,
                        "color": None,
                        "talla": talla,
                        "reason": "Ambiguo: múltiples colores para esa talla",
                    }
                )
                continue
            pending.append(
                {
                    "raw": raw,
                    "qty": qty,
                    "ref": ref,
                    "color": None,
                    "talla": talla,
                    "reason": "No se encontró ref+talla en catálogo",
                }
            )
            continue

        hits = idx_ref.get(ref, [])
        if len(hits) == 1:
            matched.append({**hits[0], "Cantidad": qty, "match_level": "ref"})
            continue
        if len(hits) > 1:
            pending.append(
                {
                    "raw": raw,
                    "qty": qty,
                    "ref": ref,
                    "color": None,
                    "talla": None,
                    "reason": "Ambiguo: referencia con múltiples variantes (resolver en grid)",
                }
            )
            continue

        pending.append(
            {
                "raw": raw,
                "qty": qty,
                "ref": ref,
                "color": None,
                "talla": None,
                "reason": "Referencia no encontrada en catálogo",
            }
        )

    return matched, pending


def add_to_cart(cart: Dict[str, dict], variant: dict, qty: int):
    ean = norm_str(variant.get("EAN", ""))
    if not ean:
        return
    qty = int(qty)
    if qty == 0:
        return

    if ean not in cart:
        cart[ean] = {
            "EAN": ean,
            "Ref": norm_ref(variant.get("Referencia", "")),
            "Nom": norm_str(variant.get("Nombre", "")),
            "Col": norm_color(variant.get("Color", "")),
            "Tal": norm_talla(variant.get("Talla", "")),
            "Cantidad": 0,
        }
    cart[ean]["Cantidad"] = int(cart[ean]["Cantidad"]) + qty
    if cart[ean]["Cantidad"] <= 0:
        cart.pop(ean, None)


def cart_to_df(cart: Dict[str, dict]) -> pd.DataFrame:
    if not cart:
        return pd.DataFrame(columns=["EAN", "Ref", "Nom", "Col", "Tal", "Cantidad"])
    df = pd.DataFrame(list(cart.values()))
    return df[["EAN", "Ref", "Nom", "Col", "Tal", "Cantidad"]].sort_values(["Ref", "Col", "Tal"])


def merge_carts(a: Dict[str, dict], b: Dict[str, dict]) -> Dict[str, dict]:
    out = {}
    for src in (a, b):
        for ean, it in src.items():
            if ean not in out:
                out[ean] = dict(it)
            else:
                out[ean]["Cantidad"] = int(out[ean]["Cantidad"]) + int(it.get("Cantidad", 0))
    out = {k: v for k, v in out.items() if int(v.get("Cantidad", 0)) > 0}
    return out


def build_search_blob(cat: pd.DataFrame) -> pd.Series:
    blob = (
        cat["EAN"].astype(str).fillna("")
        + " "
        + cat["Referencia"].astype(str).fillna("")
        + " "
        + cat["Nombre"].astype(str).fillna("")
        + " "
        + cat["Color"].astype(str).fillna("")
        + " "
        + cat["Talla"].astype(str).fillna("")
    ).str.lower()
    return blob


def export_to_template_xlsx(
    df_lines: pd.DataFrame,
    fecha: date,
    origen: str,
    destino: str,
    observaciones: str,
    template_bytes: bytes,
) -> bytes:
    if openpyxl is None:
        raise RuntimeError("Falta openpyxl. Añádelo a requirements.txt")

    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    ws = wb["Fichero ejemplo"]

    # Borra contenido previo desde fila 2
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    for i, r in df_lines.iterrows():
        row_idx = 2 + i
        ws.cell(row=row_idx, column=1).value = fecha.strftime("%Y-%m-%d")
        ws.cell(row=row_idx, column=2).value = origen
        ws.cell(row=row_idx, column=3).value = destino
        ws.cell(row=row_idx, column=4).value = observaciones
        ws.cell(row=row_idx, column=5).value = str(r["EAN"])
        ws.cell(row=row_idx, column=6).value = int(r["Cantidad"])

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ============================================================
# STATE INIT
# ============================================================

def init_state():
    st.session_state.setdefault("cat_loaded", False)
    st.session_state.setdefault("catalog_df", None)
    st.session_state.setdefault("search_blob", None)

    st.session_state.setdefault("origen", ORIGIN_OPTIONS[0])
    st.session_state.setdefault("destino", DEST_OPTIONS[0])
    st.session_state.setdefault("fecha", date.today())
    st.session_state.setdefault("ref_peticion", "")

    st.session_state.setdefault("carrito_import", {})  # EAN -> item dict
    st.session_state.setdefault("carrito_manual", {})  # EAN -> item dict

    st.session_state.setdefault("pending_rows", [])  # list dict
    st.session_state.setdefault("last_import_stats", None)

    st.session_state.setdefault("selected_ref", "")
    st.session_state.setdefault("search_query", "")


init_state()


# ============================================================
# SIDEBAR: DATA LOADING
# ============================================================

st.sidebar.markdown("### Datos")
st.sidebar.caption("Catálogo y plantilla de exportación")

catalog_file = st.sidebar.file_uploader("Catálogo (XLSX/CSV)", type=["xlsx", "xls", "csv"], key="u_catalog")
template_file = st.sidebar.file_uploader("Plantilla pedido (XLSX)", type=["xlsx"], key="u_template")


def load_catalog():
    if catalog_file is not None:
        if catalog_file.name.lower().endswith(".csv"):
            df = pd.read_csv(catalog_file)
        else:
            df = pd.read_excel(catalog_file)
        return df
    try:
        return pd.read_excel(DEFAULT_CATALOG_PATH)
    except Exception:
        return None


def load_template_bytes() -> Optional[bytes]:
    if template_file is not None:
        return template_file.getvalue()
    try:
        with open(DEFAULT_TEMPLATE_PATH, "rb") as f:
            return f.read()
    except Exception:
        return None


if st.sidebar.button("Cargar / recargar catálogo", use_container_width=True):
    df = load_catalog()
    if df is None or df.empty:
        st.sidebar.error("No se ha podido cargar el catálogo. Sube un XLSX/CSV válido.")
    else:
        needed = {"EAN", "Referencia", "Nombre", "Color", "Talla"}
        missing = needed - set(df.columns)
        if missing:
            st.sidebar.error(f"Faltan columnas en catálogo: {', '.join(sorted(missing))}")
        else:
            df = df.copy()
            df["EAN"] = df["EAN"].astype(str).str.strip()
            df["Referencia"] = df["Referencia"].map(norm_ref)
            df["Nombre"] = df["Nombre"].map(norm_str)
            df["Color"] = df["Color"].map(norm_color)
            df["Talla"] = df["Talla"].map(norm_talla)

            st.session_state.catalog_df = df
            st.session_state.search_blob = build_search_blob(df)
            st.session_state.cat_loaded = True
            st.sidebar.success(f"Catálogo cargado: {len(df):,} filas")

# Autocarga
if not st.session_state.cat_loaded and st.session_state.catalog_df is None:
    df_auto = load_catalog()
    if df_auto is not None and not df_auto.empty:
        needed = {"EAN", "Referencia", "Nombre", "Color", "Talla"}
        if needed.issubset(set(df_auto.columns)):
            df_auto = df_auto.copy()
            df_auto["EAN"] = df_auto["EAN"].astype(str).str.strip()
            df_auto["Referencia"] = df_auto["Referencia"].map(norm_ref)
            df_auto["Nombre"] = df_auto["Nombre"].map(norm_str)
            df_auto["Color"] = df_auto["Color"].map(norm_color)
            df_auto["Talla"] = df_auto["Talla"].map(norm_talla)
            st.session_state.catalog_df = df_auto
            st.session_state.search_blob = build_search_blob(df_auto)
            st.session_state.cat_loaded = True


# ============================================================
# HEADER
# ============================================================

st.markdown("# Peticiones de almacenes")
st.markdown(
    "<div class='small'>Importa una petición, revisa en dos carritos (importado/manual) y exporta a plantilla.</div>",
    unsafe_allow_html=True,
)

if not st.session_state.cat_loaded:
    st.warning("Carga el catálogo desde la barra lateral para empezar.")
    st.stop()

cat = st.session_state.catalog_df

# Índices de catálogo
idx_exact, idx_ref_color, idx_ref_talla, idx_ref = build_catalog_indexes(cat)


# ============================================================
# STEP 1: Datos del pedido
# ============================================================

st.markdown("## Datos del pedido")
c1, c2, c3, c4 = st.columns([1.2, 1.2, 1.2, 2.0])

with c1:
    st.session_state.fecha = st.date_input("Fecha", value=st.session_state.fecha)

with c2:
    st.session_state.origen = st.selectbox(
        "Almacén de origen",
        ORIGIN_OPTIONS,
        index=ORIGIN_OPTIONS.index(st.session_state.origen) if st.session_state.origen in ORIGIN_OPTIONS else 0,
    )

with c3:
    st.session_state.destino = st.selectbox(
        "Almacén de destino",
        DEST_OPTIONS,
        index=DEST_OPTIONS.index(st.session_state.destino) if st.session_state.destino in DEST_OPTIONS else 0,
    )

with c4:
    st.session_state.ref_peticion = st.text_input(
        "Referencia de la petición (se exporta en Observaciones)",
        value=st.session_state.ref_peticion,
        placeholder="Ej: PET-2026-02-08-MARBELLA",
    )

st.markdown("<hr/>", unsafe_allow_html=True)


# ============================================================
# STEP 2: Importación de petición Excel -> carrito_import
# ============================================================

st.markdown("## Importar petición (Excel) → Carrito importado")
st.caption("Se ignoran filas sin info útil. Lo ambiguo o no encontrado queda en Pendientes.")

imp_c1, imp_c2 = st.columns([2.2, 1.0])

with imp_c1:
    petition_file = st.file_uploader("Excel de petición", type=["xlsx", "xls"], key="u_petition")

with imp_c2:
    colA, colB = st.columns(2)
    with colA:
        if st.button("Vaciar carrito importado", use_container_width=True):
            st.session_state.carrito_import = {}
            st.session_state.pending_rows = []
            st.session_state.last_import_stats = None
    with colB:
        if st.button("Vaciar pendientes", use_container_width=True):
            st.session_state.pending_rows = []

if petition_file is not None and st.button("Procesar importación", type="primary"):
    pet_df = read_petition_excel(petition_file.getvalue())
    matched, pending = match_petition_to_catalog(pet_df, idx_exact, idx_ref_color, idx_ref_talla, idx_ref)

    added_lines = 0
    for m in matched:
        add_to_cart(st.session_state.carrito_import, m, int(m["Cantidad"]))
        added_lines += 1

    st.session_state.pending_rows = pending
    st.session_state.last_import_stats = {
        "matched_lines": len(matched),
        "pending_lines": len(pending),
        "added_lines": added_lines,
    }

if st.session_state.last_import_stats:
    s = st.session_state.last_import_stats
    m1, m2, m3 = st.columns(3)
    m1.metric("Líneas matcheadas", s["matched_lines"])
    m2.metric("Pendientes", s["pending_lines"])
    m3.metric("Líneas añadidas", s["added_lines"])

if st.session_state.pending_rows:
    with st.expander(f"Pendientes ({len(st.session_state.pending_rows)})", expanded=False):
        st.dataframe(pd.DataFrame(st.session_state.pending_rows), use_container_width=True)

st.markdown("<hr/>", unsafe_allow_html=True)


# ============================================================
# STEP 3: Búsqueda + Grid Color x Talla -> carrito_manual
# ============================================================

st.markdown("## Añadir manualmente → Carrito manual")
st.caption("Busca por referencia, nombre o EAN. Luego ajusta con grid Color×Talla (+ / -).")

search_c1, search_c2, search_c3 = st.columns([2.2, 1.2, 1.2])

with search_c1:
    st.session_state.search_query = st.text_input(
        "Buscar (ref / nombre / EAN / color / talla)",
        value=st.session_state.search_query,
        placeholder="Ej: 'sabhy', '248202', 'negro', '8445...'",
    )

with search_c2:
    show_limit = st.selectbox("Máx resultados", [20, 50, 100, 200], index=1)

with search_c3:
    if st.button("Vaciar carrito manual", use_container_width=True):
        st.session_state.carrito_manual = {}

q = (st.session_state.search_query or "").strip().lower()
if q:
    mask = st.session_state.search_blob.str.contains(re.escape(q), na=False)
    hits = cat.loc[mask, ["Referencia", "Nombre", "Color", "Talla", "EAN"]].copy()
else:
    hits = cat.loc[:, ["Referencia", "Nombre", "Color", "Talla", "EAN"]].head(0)

hits = hits.drop_duplicates(subset=["Referencia", "Nombre"]).head(show_limit)

ref_options = hits["Referencia"].dropna().astype(str).unique().tolist()
ref_options = sorted(ref_options)

if ref_options:
    sel_ref = st.selectbox(
        "Selecciona referencia",
        options=[""] + ref_options,
        index=0 if not st.session_state.selected_ref else (ref_options.index(st.session_state.selected_ref) + 1 if st.session_state.selected_ref in ref_options else 0),
    )
    st.session_state.selected_ref = sel_ref or ""
else:
    if q:
        st.info("No hay resultados para esa búsqueda.")
    else:
        st.info("Escribe una búsqueda para empezar.")

if st.session_state.selected_ref:
    ref = st.session_state.selected_ref
    ref_df = cat[cat["Referencia"] == ref].copy()

    if ref_df.empty:
        st.warning("No se encontraron variantes para esa referencia en catálogo.")
    else:
        nombre = ref_df["Nombre"].iloc[0] if "Nombre" in ref_df.columns else ""
        st.markdown(
            f"<div class='card'><div><span class='badge'>REF</span> "
            f"<span class='mono'>{ref}</span></div>"
            f"<div style='margin-top:6px; font-weight:600;'>{nombre}</div>"
            f"<div class='small' style='margin-top:4px;'>Ajusta cantidades por color y talla. "
            f"Se añaden al <b>carrito manual</b>.</div></div>",
            unsafe_allow_html=True,
        )

        colors = sorted(ref_df["Color"].dropna().astype(str).unique().tolist())
        tallas = sorted(ref_df["Talla"].dropna().astype(str).unique().tolist(), key=lambda x: (len(x), x))

        var_map: Dict[Tuple[str, str], dict] = {}
        for _, r in ref_df.iterrows():
            var_map[(str(r["Color"]), str(r["Talla"]))] = {
                "EAN": str(r["EAN"]),
                "Referencia": str(r["Referencia"]),
                "Nombre": str(r["Nombre"]),
                "Color": str(r["Color"]),
                "Talla": str(r["Talla"]),
            }

        st.markdown("<div class='gridwrap'>", unsafe_allow_html=True)

        header_cols = st.columns([1.2] + [1.0] * len(colors))
        header_cols[0].markdown("**Talla \\ Color**")
        for j, col in enumerate(colors, start=1):
            header_cols[j].markdown(f"**{col}**")

        for talla in tallas:
            row_cols = st.columns([1.2] + [1.0] * len(colors))
            row_cols[0].markdown(f"**{talla}**")

            for j, col in enumerate(colors, start=1):
                variant = var_map.get((col, talla))

                if not variant:
                    row_cols[j].markdown("<div class='cellqty small'>—</div>", unsafe_allow_html=True)
                    continue

                ean = variant["EAN"]
                current_qty = int(st.session_state.carrito_manual.get(ean, {}).get("Cantidad", 0))

                b1, b2, b3 = row_cols[j].columns([1, 1, 1])
                with b1:
                    if st.button("−", key=f"minus_{ref}_{col}_{talla}"):
                        add_to_cart(st.session_state.carrito_manual, variant, -1)
                        st.rerun()
                with b2:
                    st.markdown(f"<div class='cellqty'>{current_qty}</div>", unsafe_allow_html=True)
                with b3:
                    if st.button("+", key=f"plus_{ref}_{col}_{talla}"):
                        add_to_cart(st.session_state.carrito_manual, variant, +1)
                        st.rerun()

        st.markdown("</div>", unsafe_allow_html=True)

st.markdown("<hr/>", unsafe_allow_html=True)


# ============================================================
# STEP 4: Revisión (2 carritos + fusión)
# ============================================================

st.markdown("## Revisión final")
cA, cB, cC = st.columns(3)

car_imp_df = cart_to_df(st.session_state.carrito_import)
car_man_df = cart_to_df(st.session_state.carrito_manual)
car_merge = merge_carts(st.session_state.carrito_import, st.session_state.carrito_manual)

car_merge_df = cart_to_df(car_merge)[["EAN", "Ref", "Nom", "Col", "Tal", "Cantidad"]].rename(
    columns={"Ref": "Referencia", "Nom": "Nombre", "Col": "Color", "Tal": "Talla"}
)

def apply_edited_cart(edited: pd.DataFrame) -> Dict[str, dict]:
    new_cart: Dict[str, dict] = {}
    for _, r in edited.iterrows():
        ean = str(r["EAN"])
        qty = int(r["Cantidad"])
        if qty > 0:
            new_cart[ean] = {
                "EAN": ean,
                "Ref": str(r["Referencia"]),
                "Nom": str(r["Nombre"]),
                "Col": str(r["Color"]),
                "Tal": str(r["Talla"]),
                "Cantidad": qty,
            }
    return new_cart


with cA:
    st.markdown("<div class='card'><b>Carrito importado</b></div>", unsafe_allow_html=True)
    st.caption(f"{len(car_imp_df)} líneas")
    if car_imp_df.empty:
        st.info("Vacío.")
    else:
        edit_df = car_imp_df.rename(columns={"Ref": "Referencia", "Nom": "Nombre", "Col": "Color", "Tal": "Talla"})
        edited = st.data_editor(
            edit_df,
            use_container_width=True,
            hide_index=True,
            disabled=["EAN", "Referencia", "Nombre", "Color", "Talla"],
        )
        st.session_state.carrito_import = apply_edited_cart(edited)

with cB:
    st.markdown("<div class='card'><b>Carrito manual</b></div>", unsafe_allow_html=True)
    st.caption(f"{len(car_man_df)} líneas")
    if car_man_df.empty:
        st.info("Vacío.")
    else:
        edit_df = car_man_df.rename(columns={"Ref": "Referencia", "Nom": "Nombre", "Col": "Color", "Tal": "Talla"})
        edited = st.data_editor(
            edit_df,
            use_container_width=True,
            hide_index=True,
            disabled=["EAN", "Referencia", "Nombre", "Color", "Talla"],
        )
        st.session_state.carrito_manual = apply_edited_cart(edited)

with cC:
    st.markdown("<div class='card'><b>Fusión (se exporta)</b></div>", unsafe_allow_html=True)
    st.caption(f"{len(car_merge_df)} líneas")
    if car_merge_df.empty:
        st.info("No hay líneas para exportar.")
    else:
        st.dataframe(car_merge_df, use_container_width=True, hide_index=True)

st.markdown("<hr/>", unsafe_allow_html=True)


# ============================================================
# STEP 5: Export
# ============================================================

st.markdown("## Exportación")

tpl_bytes = load_template_bytes()
if tpl_bytes is None:
    st.warning("Sube la plantilla 'plantilla_pedido.xlsx' en la barra lateral (o deja ese fichero en el directorio).")

if openpyxl is None:
    st.error("No se puede exportar a plantilla: falta la dependencia 'openpyxl'.")
    st.stop()

if car_merge_df.empty:
    st.info("No hay líneas para exportar.")
    st.stop()

export_lines = car_merge_df[["EAN", "Cantidad"]].copy()
export_lines["Cantidad"] = pd.to_numeric(export_lines["Cantidad"], errors="coerce").fillna(0).astype(int)
export_lines = export_lines[export_lines["Cantidad"] > 0].reset_index(drop=True)

colX, colY, colZ = st.columns([1.2, 1.2, 2.0])
with colX:
    filename = st.text_input(
        "Nombre archivo",
        value=f"pedido_{st.session_state.origen}_a_{st.session_state.destino}_{st.session_state.fecha.isoformat()}.xlsx".replace(" ", "_"),
    )
with colY:
    st.caption("Observaciones (export):")
    st.code(st.session_state.ref_peticion or "-", language="text")
with colZ:
    st.caption("Resumen:")
    st.write(
        f"- Fecha: **{st.session_state.fecha.isoformat()}**\n"
        f"- Origen: **{st.session_state.origen}**\n"
        f"- Destino: **{st.session_state.destino}**\n"
        f"- Líneas: **{len(export_lines)}**"
    )

if tpl_bytes is not None:
    try:
        out_bytes = export_to_template_xlsx(
            df_lines=export_lines,
            fecha=st.session_state.fecha,
            origen=st.session_state.origen,
            destino=st.session_state.destino,
            observaciones=st.session_state.ref_peticion or "",
            template_bytes=tpl_bytes,
        )
        st.download_button(
            "Descargar pedido (XLSX)",
            data=out_bytes,
            file_name=filename if filename.lower().endswith(".xlsx") else f"{filename}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
    except Exception as e:
        st.error(f"Error exportando: {e}")

st.markdown(
    "<div class='small'>Tip: si el importador deja líneas en Pendientes por ambigüedad, "
    "resuélvelas con el grid seleccionando la referencia y ajustando color/talla.</div>",
    unsafe_allow_html=True,
)
