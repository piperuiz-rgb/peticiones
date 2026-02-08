## utils.py
from __future__ import annotations

import io
import re
from datetime import date
from typing import Dict, Tuple, Optional, List

import pandas as pd
import streamlit as st

try:
    import openpyxl
except Exception:
    openpyxl = None

# -----------------------------
# Config & constants
# -----------------------------
ORIGIN_OPTIONS = [
    "PET Almacén Badalona",
    "ALM-CENTRAL",
]
DEST_OPTIONS = [
    "PET T002 Marbella",
    "ALM-TIENDA",
]

DEFAULT_CATALOG_PATH = "catalogue.xlsx"
DEFAULT_TEMPLATE_PATH = "plantilla_pedido.xlsx"

TALLA_MAP = {
    "XXS": "XXS",
    "XS": "XS",
    "S": "S",
    "M": "M",
    "L": "L",
    "XL": "XL",
    "XXL": "XXL",
    "XXXL": "XXXL",
}
TALLA_REGEX = re.compile(r"^\\s*(XXS|XS|S|M|L|XL|XXL|XXXL|[0-9]{2,3}|[0-9]{1,2}[A-Z]?)\\s*$", re.I)
REF_BRACKET_REGEX = re.compile(r"\[(?P<ref>[^\]]+)\]")
ATTR_PAREN_REGEX = re.compile(r"\((?P<attrs>[^)]+)\)\s*$")

def ensure_style():
    st.markdown(
        """
<style>
:root { --accent: #111; }
.block-container { padding-top: 1.2rem; padding-bottom: 2.5rem; }
h1, h2, h3 { letter-spacing: -0.02em; }
div[data-testid="stSidebar"] { border-right: 1px solid #eee; }
hr { margin: 1.2rem 0; }
.card { border: 1px solid #eee; border-radius: 14px; padding: 14px 16px; background: #fff; }
.badge { display:inline-block; padding:2px 10px; border-radius:999px; border:1px solid #eaeaea; font-size:12px; background:#fafafa; }
.small { font-size: 12px; color: #666; }
.mono { font-family: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Liberation Mono", "Courier New", monospace; }
.gridwrap { border: 1px solid #eee; border-radius: 14px; padding: 10px 12px; background: #fff; overflow-x: auto; }
.cellqty { text-align: center; font-weight: 700; padding-top: 6px; padding-bottom: 6px; }
button[kind="secondary"] { font-size: 18px; font-weight: 700; }
</style>
""",
        unsafe_allow_html=True,
    )


def init_state():
    st.session_state.setdefault("cat_loaded", False)
    st.session_state.setdefault("catalog_df", None)
    st.session_state.setdefault("search_blob", None)

    st.session_state.setdefault("origen", ORIGIN_OPTIONS[0])
    st.session_state.setdefault("destino", DEST_OPTIONS[0])
    st.session_state.setdefault("fecha", date.today())
    st.session_state.setdefault("ref_peticion", "")

    st.session_state.setdefault("carrito_import", {})
    st.session_state.setdefault("carrito_manual", {})

    st.session_state.setdefault("pending_rows", [])
    st.session_state.setdefault("last_import_stats", None)

    st.session_state.setdefault("selected_ref", "")
    st.session_state.setdefault("search_query", "")


def norm_str(x: object) -> str:
    if x is None:
        return ""
    return str(x).strip()


def norm_ref(x: object) -> str:
    return norm_str(x)


def norm_color(x: object) -> str:
    return norm_str(x)


def norm_talla(x: object) -> str:
    s = norm_str(x)
    up = s.upper()
    return TALLA_MAP.get(up, s)


def looks_like_talla(s: str) -> bool:
    if not s:
        return False
    return bool(TALLA_REGEX.match(s.strip()))


def build_search_blob(cat: pd.DataFrame) -> pd.Series:
    blob = (
        cat["EAN"].astype(str).fillna("")
        + " "
        + cat["Referencia"].astype(str).fillna("")
        + " "
        + cat["Nombre"].astype(str).fillna("")
        + " "
        + cat["Color"].astype(str).fillna("")
        + " "
        + cat["Talla"].astype(str).fillna("")
    ).str.lower()
    return blob


def load_catalog_from_sidebar():
    st.sidebar.markdown("### Datos")
    st.sidebar.caption("Catálogo y plantilla de exportación")

    catalog_file = st.sidebar.file_uploader("Catálogo (XLSX/CSV)", type=["xlsx", "xls", "csv"], key="u_catalog")
    template_file = st.sidebar.file_uploader("Plantilla pedido (XLSX)", type=["xlsx"], key="u_template")

    def load_catalog():
        if catalog_file is not None:
            if catalog_file.name.lower().endswith(".csv"):
                return pd.read_csv(catalog_file)
            return pd.read_excel(catalog_file)
        try:
            return pd.read_excel(DEFAULT_CATALOG_PATH)
        except Exception:
            return None

    def load_template_bytes() -> Optional[bytes]:
        if template_file is not None:
            return template_file.getvalue()
        try:
            with open(DEFAULT_TEMPLATE_PATH, "rb") as f:
                return f.read()
        except Exception:
            return None

    if st.sidebar.button("Cargar / recargar catálogo", use_container_width=True):
        df = load_catalog()
        if df is None or df.empty:
            st.sidebar.error("No se ha podido cargar el catálogo. Sube un XLSX/CSV válido.")
        else:
            needed = {"EAN", "Referencia", "Nombre", "Color", "Talla"}
            missing = needed - set(df.columns)
            if missing:
                st.sidebar.error(f"Faltan columnas en catálogo: {', '.join(sorted(missing))}")
            else:
                df = df.copy()
                df["EAN"] = df["EAN"].astype(str).str.strip()
                df["Referencia"] = df["Referencia"].map(norm_ref)
                df["Nombre"] = df["Nombre"].map(norm_str)
                df["Color"] = df["Color"].map(norm_color)
                df["Talla"] = df["Talla"].map(norm_talla)
                st.session_state.catalog_df = df
                st.session_state.search_blob = build_search_blob(df)
                st.session_state.cat_loaded = True
                st.sidebar.success(f"Catálogo cargado: {len(df):,} filas")

    # Autocarga
    if not st.session_state.cat_loaded and st.session_state.catalog_df is None:
        df_auto = load_catalog()
        if df_auto is not None and not df_auto.empty:
            needed = {"EAN", "Referencia", "Nombre", "Color", "Talla"}
            if needed.issubset(set(df_auto.columns)):
                df_auto = df_auto.copy()
                df_auto["EAN"] = df_auto["EAN"].astype(str).str.strip()
                df_auto["Referencia"] = df_auto["Referencia"].map(norm_ref)
                df_auto["Nombre"] = df_auto["Nombre"].map(norm_str)
                df_auto["Color"] = df_auto["Color"].map(norm_color)
                df_auto["Talla"] = df_auto["Talla"].map(norm_talla)
                st.session_state.catalog_df = df_auto
                st.session_state.search_blob = build_search_blob(df_auto)
                st.session_state.cat_loaded = True

    return load_template_bytes()


def build_catalog_indexes(cat: pd.DataFrame):
    idx_exact: Dict[Tuple[str, str, str], dict] = {}
    idx_ref_color: Dict[Tuple[str, str], List[dict]] = {}
    idx_ref_talla: Dict[Tuple[str, str], List[dict]] = {}
    idx_ref: Dict[str, List[dict]] = {}

    for _, r in cat.iterrows():
        ref = norm_ref(r.get("Referencia", ""))
        color = norm_color(r.get("Color", ""))
        talla = norm_talla(r.get("Talla", ""))
        ean = norm_str(r.get("EAN", ""))
        nombre = norm_str(r.get("Nombre", ""))

        row = {"EAN": ean, "Referencia": ref, "Color": color, "Talla": talla, "Nombre": nombre}
        idx_exact[(ref, color, talla)] = row
        idx_ref_color.setdefault((ref, color), []).append(row)
        idx_ref_talla.setdefault((ref, talla), []).append(row)
        idx_ref.setdefault(ref, []).append(row)

    return idx_exact, idx_ref_color, idx_ref_talla, idx_ref


def parse_petition_line(raw: str) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    if not raw:
        return None, None, None
    raw = str(raw).strip()
    mref = REF_BRACKET_REGEX.search(raw)
    if not mref:
        return None, None, None
    ref = norm_ref(mref.group("ref"))

    mattr = ATTR_PAREN_REGEX.search(raw)
    if not mattr:
        return ref, None, None

    attrs_txt = mattr.group("attrs")
    parts = [p.strip() for p in attrs_txt.split(",") if p.strip()]

    color = None
    talla = None
    if len(parts) >= 2:
        a, b = parts[0], parts[1]
        if looks_like_talla(a) and not looks_like_talla(b):
            talla = norm_talla(a); color = norm_color(b)
        elif looks_like_talla(b) and not looks_like_talla(a):
            color = norm_color(a); talla = norm_talla(b)
        else:
            color = norm_color(a); talla = norm_talla(b)
    elif len(parts) == 1:
        a = parts[0]
        if looks_like_talla(a):
            talla = norm_talla(a)
        else:
            color = norm_color(a)

    return ref, color, talla


def detect_qty_column(df: pd.DataFrame) -> str:
    cols = list(df.columns)
    if len(cols) >= 3:
        c2 = cols[2]
        s2 = pd.to_numeric(df[c2], errors="coerce").fillna(0).sum()
        if s2 > 0:
            return c2
    if len(cols) >= 2:
        return cols[1]
    return cols[0]


def read_petition_excel(file_bytes: bytes) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(file_bytes))
    cols = list(df.columns)
    if not cols:
        return pd.DataFrame(columns=["raw", "qty"])
    raw_col = cols[0]
    qty_col = detect_qty_column(df)
    return pd.DataFrame(
        {"raw": df[raw_col], "qty": pd.to_numeric(df[qty_col], errors="coerce").fillna(0).astype(int)}
    )


def match_petition_to_catalog(petition_df: pd.DataFrame, idx_exact, idx_ref_color, idx_ref_talla, idx_ref):
    matched = []
    pending = []
    for _, r in petition_df.iterrows():
        raw = norm_str(r.get("raw", ""))
        qty = int(r.get("qty", 0))
        if qty <= 0:
            continue
        if "[" not in raw:
            continue

        ref, color, talla = parse_petition_line(raw)
        if not ref:
            continue

        if color and talla:
            hit = idx_exact.get((ref, color, talla))
            if hit:
                matched.append({**hit, "Cantidad": qty, "match_level": "exact"})
            else:
                pending.append({"raw": raw, "qty": qty, "ref": ref, "color": color, "talla": talla,
                                "reason": "No existe esa variante exacta en catálogo"})
            continue

        if color and not talla:
            hits = idx_ref_color.get((ref, color), [])
            if len(hits) == 1:
                matched.append({**hits[0], "Cantidad": qty, "match_level": "ref+color"})
            elif len(hits) > 1:
                pending.append({"raw": raw, "qty": qty, "ref": ref, "color": color, "talla": None,
                                "reason": "Ambiguo: múltiples tallas para ese color"})
            else:
                pending.append({"raw": raw, "qty": qty, "ref": ref, "color": color, "talla": None,
                                "reason": "No se encontró ref+color en catálogo"})
            continue

        if talla and not color:
            hits = idx_ref_talla.get((ref, talla), [])
            if len(hits) == 1:
                matched.append({**hits[0], "Cantidad": qty, "match_level": "ref+talla"})
            elif len(hits) > 1:
                pending.append({"raw": raw, "qty": qty, "ref": ref, "color": None, "talla": talla,
                                "reason": "Ambiguo: múltiples colores para esa talla"})
            else:
                pending.append({"raw": raw, "qty": qty, "ref": ref, "color": None, "talla": talla,
                                "reason": "No se encontró ref+talla en catálogo"})
            continue

        hits = idx_ref.get(ref, [])
        if len(hits) == 1:
            matched.append({**hits[0], "Cantidad": qty, "match_level": "ref"})
        elif len(hits) > 1:
            pending.append({"raw": raw, "qty": qty, "ref": ref, "color": None, "talla": None,
                            "reason": "Ambiguo: referencia con múltiples variantes (resolver en grid)"})
        else:
            pending.append({"raw": raw, "qty": qty, "ref": ref, "color": None, "talla": None,
                            "reason": "Referencia no encontrada en catálogo"})
    return matched, pending


def add_to_cart(cart: Dict[str, dict], variant: dict, qty: int):
    ean = norm_str(variant.get("EAN", ""))
    if not ean:
        return
    qty = int(qty)
    if qty == 0:
        return

    if ean not in cart:
        cart[ean] = {
            "EAN": ean,
            "Ref": norm_ref(variant.get("Referencia", "")),
            "Nom": norm_str(variant.get("Nombre", "")),
            "Col": norm_color(variant.get("Color", "")),
            "Tal": norm_talla(variant.get("Talla", "")),
            "Cantidad": 0,
        }
    cart[ean]["Cantidad"] = int(cart[ean]["Cantidad"]) + qty
    if cart[ean]["Cantidad"] <= 0:
        cart.pop(ean, None)


def cart_to_df(cart: Dict[str, dict]) -> pd.DataFrame:
    if not cart:
        return pd.DataFrame(columns=["EAN", "Ref", "Nom", "Col", "Tal", "Cantidad"])
    df = pd.DataFrame(list(cart.values()))
    return df[["EAN", "Ref", "Nom", "Col", "Tal", "Cantidad"]].sort_values(["Ref", "Col", "Tal"])


def merge_carts(a: Dict[str, dict], b: Dict[str, dict]) -> Dict[str, dict]:
    out = {}
    for src in (a, b):
        for ean, it in src.items():
            if ean not in out:
                out[ean] = dict(it)
            else:
                out[ean]["Cantidad"] = int(out[ean]["Cantidad"]) + int(it.get("Cantidad", 0))
    return {k: v for k, v in out.items() if int(v.get("Cantidad", 0)) > 0}


def export_to_template_xlsx(df_lines: pd.DataFrame, fecha: date, origen: str, destino: str, observaciones: str, template_bytes: bytes) -> bytes:
    if openpyxl is None:
        raise RuntimeError("Falta openpyxl. Añádelo a requirements.txt")
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    ws = wb["Fichero ejemplo"]
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)
    for i, r in df_lines.iterrows():
        row_idx = 2 + i
        ws.cell(row=row_idx, column=1).value = fecha.strftime("%Y-%m-%d")
        ws.cell(row=row_idx, column=2).value = origen
        ws.cell(row=row_idx, column=3).value = destino
        ws.cell(row=row_idx, column=4).value = observaciones
        ws.cell(row=row_idx, column=5).value = str(r["EAN"])
        ws.cell(row=row_idx, column=6).value = int(r["Cantidad"])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
