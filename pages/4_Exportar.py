# pages/4_Exportar.py
import io
import re
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from utils import init_state, ensure_style, load_repo_data, merge_carts

st.set_page_config(page_title="Exportar", page_icon="📦", layout="wide")
ensure_style()
init_state()
load_repo_data()

st.markdown("# 4 · Exportar pedido")
st.markdown(
    "<div class='small'>Genera el Excel final usando la plantilla del repositorio.</div>",
    unsafe_allow_html=True,
)

# -----------------------------
# Helpers
# -----------------------------
def _safe(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9._ -]+", "_", (s or "").strip())

def _find_cell_by_label(ws, label_text: str):
    """Busca una celda que contenga el texto label_text y devuelve (row, col) de la celda a la derecha (valor)."""
    target = (label_text or "").strip().lower()
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80), values_only=False):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.strip().lower() == target:
                # valor suele ir a la derecha
                return (cell.row, cell.column + 1)
    return None

def _find_header_row(ws, required_headers):
    """
    Encuentra una fila de cabecera que contenga varios encabezados.
    Devuelve dict header->col y row_index.
    """
    req = [h.strip().lower() for h in required_headers]
    for r in range(1, min(ws.max_row, 120) + 1):
        row_vals = []
        for c in range(1, min(ws.max_column, 80) + 1):
            v = ws.cell(r, c).value
            row_vals.append((c, v.strip().lower() if isinstance(v, str) else v))
        # mapa de strings
        colmap = {}
        for c, v in row_vals:
            if isinstance(v, str):
                colmap[v] = c
        hits = sum(1 for h in req if h in colmap)
        if hits >= max(2, min(3, len(req))):  # heurística: al menos 2-3 headers
            return colmap, r
    return None, None

# -----------------------------
# Datos
# -----------------------------
merged = merge_carts(st.session_state.get("carrito_import", {}), st.session_state.get("carrito_manual", {}))

if not merged:
    st.warning("No hay líneas en el pedido. Vuelve a Revisión y añade prendas.")
    st.page_link("pages/3_Revision_final.py", label="← Volver a 3 · Revisión final", use_container_width=True)
    st.stop()

if st.session_state.get("tpl_bytes") is None:
    st.error("No se encontró **plantilla_pedido.xlsx** en el repositorio. No se puede exportar.")
    st.stop()

# Cabecera (si ya estás usando PET en session_state, aquí irá PET)
origen_txt = st.session_state.get("origen", "")
destino_txt = st.session_state.get("destino", "")
ref_txt = st.session_state.get("ref_peticion", "")

if origen_txt == destino_txt:
    st.error("Origen y destino no pueden coincidir.")
    st.stop()

# Nombre de archivo
safe_ref = _safe(ref_txt) if ref_txt else "SIN_REF"
filename = f"{st.session_state.get('fecha'):%Y%m%d}_{safe_ref}.xlsx".replace(" ", "_")

# -----------------------------
# Cargar plantilla
# -----------------------------
wb = load_workbook(io.BytesIO(st.session_state.tpl_bytes))
ws = wb.active

# -----------------------------
# Escribir cabecera (intenta por etiquetas; si no, fallback a B2/B3/B4)
# -----------------------------
# Primero, por etiquetas exactas típicas
pos_origen = _find_cell_by_label(ws, "origen")
pos_destino = _find_cell_by_label(ws, "destino")
pos_obs = _find_cell_by_label(ws, "observaciones")

# Fallback: celdas típicas
if pos_origen:
    ws.cell(pos_origen[0], pos_origen[1]).value = origen_txt
else:
    ws["B2"].value = origen_txt

if pos_destino:
    ws.cell(pos_destino[0], pos_destino[1]).value = destino_txt
else:
    ws["B3"].value = destino_txt

if pos_obs:
    ws.cell(pos_obs[0], pos_obs[1]).value = ref_txt
else:
    ws["B4"].value = ref_txt

# -----------------------------
# Preparar líneas a exportar
# -----------------------------
rows = []
for ean, it in merged.items():
    rows.append(
        {
            "EAN": str(ean),
            "Ref": it.get("Ref", ""),
            "Nombre": it.get("Nom", ""),
            "Color": it.get("Col", ""),
            "Talla": it.get("Tal", ""),
            "Cantidad": int(it.get("Cantidad", 0) or 0),
            "Observaciones": ref_txt or "",
        }
    )
rows = [r for r in rows if r["Cantidad"] > 0]

# Orden bonito
rows.sort(key=lambda x: (x["Ref"], x["Color"], x["Talla"], x["EAN"]))

# -----------------------------
# Escribir tabla: intenta detectar cabecera; si no, empieza en fila 10
# -----------------------------
# Headers posibles según plantilla
candidate_headers = ["ean", "ref", "referencia", "nombre", "color", "talla", "cantidad", "uds", "unidades", "observaciones"]
colmap, header_row = _find_header_row(ws, candidate_headers)

if header_row:
    start_row = header_row + 1

    # Construye mapeo “nuestro campo” -> columna real detectada
    # Priorizamos cabeceras exactas si existen
    def col_for(*names):
        for n in names:
            n = n.lower()
            if n in colmap:
                return colmap[n]
        return None

    c_ean = col_for("ean", "barcode", "codbarras")
    c_ref = col_for("ref", "referencia")
    c_nom = col_for("nombre", "name", "producto")
    c_col = col_for("color", "col")
    c_tal = col_for("talla", "tal", "size")
    c_qty = col_for("cantidad", "uds", "unidades", "qty")
    c_obs = col_for("observaciones", "obs")

    # Fallback si la plantilla no tiene algunas columnas detectables
    # (ponemos un layout razonable desde la primera columna del header)
    base_col = min(colmap.values()) if colmap else 1
    if c_ean is None: c_ean = base_col + 0
    if c_ref is None: c_ref = base_col + 1
    if c_nom is None: c_nom = base_col + 2
    if c_col is None: c_col = base_col + 3
    if c_tal is None: c_tal = base_col + 4
    if c_qty is None: c_qty = base_col + 5
    if c_obs is None: c_obs = base_col + 6

else:
    # Fallback fijo
    start_row = 10
    c_ean, c_ref, c_nom, c_col, c_tal, c_qty, c_obs = 1, 2, 3, 4, 5, 6, 7

# Limpia filas antiguas (opcional y seguro: borra hasta 500 filas)
for r in range(start_row, start_row + 500):
    # si no hay nada en las columnas clave, paramos pronto
    if all(ws.cell(r, c).value in (None, "") for c in (c_ean, c_ref, c_qty)):
        break
    for c in (c_ean, c_ref, c_nom, c_col, c_tal, c_qty, c_obs):
        ws.cell(r, c).value = None

# Escribe filas
r = start_row
for item in rows:
    ws.cell(r, c_ean).value = item["EAN"]
    ws.cell(r, c_ref).value = item["Ref"]
    ws.cell(r, c_nom).value = item["Nombre"]
    ws.cell(r, c_col).value = item["Color"]
    ws.cell(r, c_tal).value = item["Talla"]
    ws.cell(r, c_qty).value = item["Cantidad"]
    ws.cell(r, c_obs).value = item["Observaciones"]
    r += 1

# -----------------------------
# Descargar
# -----------------------------
out = io.BytesIO()
wb.save(out)
out.seek(0)

st.success("Archivo listo para descargar.")
st.download_button(
    "Descargar Excel",
    data=out.getvalue(),
    file_name=filename,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)

st.markdown("<hr/>", unsafe_allow_html=True)
st.page_link("pages/3_Revision_final.py", label="← Volver a 3 · Revisión final", use_container_width=True)
